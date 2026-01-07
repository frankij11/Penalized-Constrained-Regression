"""
Export functions for diagnostic reports.

Provides HTML, PDF, and Excel export capabilities with support for
embedded plots, sample data, and model equations.
"""

import warnings
from typing import Optional, List, TYPE_CHECKING
import numpy as np

if TYPE_CHECKING:
    from .dataclasses import (
        SummaryReport, CoefficientInfo, FitStatistics, ModelSpecification,
        DataSummary, ConstraintSummary, ResidualAnalysis, SampleData, ModelEquation,
        BootstrapResults
    )


def to_excel(
    report: 'SummaryReport',
    filepath: str,
    sample_n: int = 50,
    single_sheet: bool = True
):
    """
    Export report to Excel file.

    Parameters
    ----------
    report : SummaryReport
        The summary report to export
    filepath : str
        Path to output Excel file (.xlsx).
    sample_n : int, default=50
        Number of sample rows to include. -1 for all.
    single_sheet : bool, default=True
        If True, create a single formatted sheet like HTML.
        If False, use multiple sheets (legacy behavior).
    """
    try:
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError("pandas and openpyxl are required for to_excel()")

    if single_sheet:
        _to_excel_single_sheet(report, filepath, sample_n)
    else:
        _to_excel_multi_sheet(report, filepath, sample_n)


def _to_excel_single_sheet(report: 'SummaryReport', filepath: str, sample_n: int):
    """Create a single formatted Excel sheet matching HTML report style."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Model Summary"

    # Styles
    title_font = Font(bold=True, size=16, color="2C3E50")
    section_font = Font(bold=True, size=12, color="34495E")
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    metric_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    row = 1

    # Title
    ws.cell(row=row, column=1, value="PENALIZED-CONSTRAINED REGRESSION SUMMARY").font = title_font
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    row += 2

    # Key Metrics (horizontal cards)
    ws.cell(row=row, column=1, value="R2").font = Font(bold=True)
    ws.cell(row=row, column=2, value="SPE").font = Font(bold=True)
    ws.cell(row=row, column=3, value="MAPE").font = Font(bold=True)
    ws.cell(row=row, column=4, value="RMSE").font = Font(bold=True)
    ws.cell(row=row, column=5, value="Active Constraints").font = Font(bold=True)
    for col in range(1, 6):
        ws.cell(row=row, column=col).fill = metric_fill
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 1

    ws.cell(row=row, column=1, value=f"{report.fit_stats.r2:.4f}")
    ws.cell(row=row, column=2, value=f"{report.fit_stats.spe:.2%}")
    ws.cell(row=row, column=3, value=f"{report.fit_stats.mape:.2%}")
    ws.cell(row=row, column=4, value=f"{report.fit_stats.rmse:.4f}")
    ws.cell(row=row, column=5, value=report.constraints.n_active)
    for col in range(1, 6):
        ws.cell(row=row, column=col).alignment = Alignment(horizontal='center')
    row += 2

    # Model Equation
    if report.equation is not None:
        ws.cell(row=row, column=1, value="MODEL EQUATION").font = section_font
        row += 1
        ws.cell(row=row, column=1, value=report.equation.text)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        row += 2

    # Model Specification
    ws.cell(row=row, column=1, value="MODEL SPECIFICATION").font = section_font
    row += 1
    spec_items = [
        ("Model Type", report.model_spec.model_type),
        ("Loss Function", report.model_spec.loss_function),
        ("Alpha", report.model_spec.alpha),
        ("L1 Ratio", report.model_spec.l1_ratio),
        ("Method", report.model_spec.method),
        ("Converged", report.model_spec.converged),
    ]
    if report.model_spec.fit_datetime:
        spec_items.append(("Fit Date", str(report.model_spec.fit_datetime)[:19]))
    if report.model_spec.fit_duration_seconds:
        spec_items.append(("Fit Duration", f"{report.model_spec.fit_duration_seconds:.2f}s"))

    for i, (label, value) in enumerate(spec_items):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value) if value is not None else "N/A")
        row += 1
    row += 1

    # Data Summary
    ws.cell(row=row, column=1, value="DATA SUMMARY").font = section_font
    row += 1
    data_items = [
        ("N Samples", report.data_summary.n_samples),
        ("N Features", report.data_summary.n_features),
        ("Y Mean", f"{report.data_summary.y_mean:.4f}"),
        ("Y Std", f"{report.data_summary.y_std:.4f}"),
        ("Y Range", f"[{report.data_summary.y_min:.4f}, {report.data_summary.y_max:.4f}]"),
    ]
    for label, value in data_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1
    row += 1

    # Coefficients Table
    ws.cell(row=row, column=1, value="COEFFICIENTS").font = section_font
    if report.ci_method != 'none':
        ws.cell(row=row, column=3, value=f"(95% CIs via {report.ci_method})")
    row += 1

    # Header row
    has_ci = any(c.ci_lower is not None for c in report.coefficients)
    headers = ["Parameter", "Value"]
    if has_ci:
        headers.extend(["CI Lower", "CI Upper"])
    headers.extend(["Bounds", "Status"])

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    row += 1

    # Coefficient rows
    all_coefs = list(report.coefficients)
    if report.intercept is not None:
        from .dataclasses import CoefficientInfo
        all_coefs.append(CoefficientInfo(
            name='Intercept',
            value=report.intercept.value,
            lower_bound=report.intercept.lower_bound,
            upper_bound=report.intercept.upper_bound,
            is_at_lower=report.intercept.is_at_lower,
            is_at_upper=report.intercept.is_at_upper,
            se=report.intercept.se,
            ci_lower=report.intercept.ci_lower,
            ci_upper=report.intercept.ci_upper,
        ))

    for c in all_coefs:
        col = 1
        ws.cell(row=row, column=col, value=c.name).border = thin_border
        col += 1
        ws.cell(row=row, column=col, value=f"{c.value:.6f}").border = thin_border
        col += 1
        if has_ci:
            ws.cell(row=row, column=col, value=f"{c.ci_lower:.4f}" if c.ci_lower else "-").border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=f"{c.ci_upper:.4f}" if c.ci_upper else "-").border = thin_border
            col += 1
        lb = f"{c.lower_bound:.4f}" if np.isfinite(c.lower_bound) else "-inf"
        ub = f"{c.upper_bound:.4f}" if np.isfinite(c.upper_bound) else "+inf"
        ws.cell(row=row, column=col, value=f"[{lb}, {ub}]").border = thin_border
        col += 1
        status_cell = ws.cell(row=row, column=col, value=c.bound_status)
        status_cell.border = thin_border
        if c.is_constrained:
            status_cell.font = Font(color="E74C3C", bold=True)
        else:
            status_cell.font = Font(color="27AE60")
        row += 1
    row += 1

    # Fit Statistics
    ws.cell(row=row, column=1, value="FIT STATISTICS").font = section_font
    row += 1
    fit_items = [
        ("R2", f"{report.fit_stats.r2:.4f}"),
        ("Adjusted R2", f"{report.fit_stats.adj_r2:.4f}"),
        ("SEE", f"{report.fit_stats.see:.4f}"),
        ("SPE", f"{report.fit_stats.spe:.2%}"),
        ("MAPE", f"{report.fit_stats.mape:.2%}"),
        ("RMSE", f"{report.fit_stats.rmse:.4f}"),
    ]
    if report.fit_stats.mse is not None:
        fit_items.append(("MSE", f"{report.fit_stats.mse:.4f}"))
    if report.fit_stats.mae is not None:
        fit_items.append(("MAE", f"{report.fit_stats.mae:.4f}"))
    fit_items.append(("CV", f"{report.fit_stats.cv:.2%}"))
    fit_items.append((f"GDF ({report.fit_stats.gdf_method})", f"{report.fit_stats.gdf:.1f}"))
    if report.fit_stats.aic is not None:
        fit_items.append(("AIC", f"{report.fit_stats.aic:.4f}"))
    if report.fit_stats.bic is not None:
        fit_items.append(("BIC", f"{report.fit_stats.bic:.4f}"))
    if report.fit_stats.f_statistic is not None:
        pval_str = f"{report.fit_stats.f_pvalue:.4e}" if report.fit_stats.f_pvalue else "N/A"
        fit_items.append(("F-statistic", f"{report.fit_stats.f_statistic:.4f} (p={pval_str})"))
    if report.fit_stats.durbin_watson is not None:
        fit_items.append(("Durbin-Watson", f"{report.fit_stats.durbin_watson:.4f}"))
    for label, value in fit_items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)
        row += 1
    row += 1

    # Residual Analysis
    if report.residuals is not None:
        ws.cell(row=row, column=1, value="RESIDUAL ANALYSIS").font = section_font
        row += 1
        resid_items = [
            ("Mean", f"{report.residuals.mean:.4f}"),
            ("Std", f"{report.residuals.std:.4f}"),
            ("Range", f"[{report.residuals.min:.4f}, {report.residuals.max:.4f}]"),
            ("Skewness", f"{report.residuals.skewness:.4f}"),
            ("Kurtosis", f"{report.residuals.kurtosis:.4f}"),
            ("Outliers (|r| > 2*std)", report.residuals.n_outliers),
        ]
        for label, value in resid_items:
            ws.cell(row=row, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row, column=2, value=str(value))
            row += 1
        row += 1

    # Constraints
    if report.constraints.active_constraints:
        ws.cell(row=row, column=1, value="ACTIVE CONSTRAINTS").font = section_font
        row += 1
        for name, bound_type in report.constraints.active_constraints:
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=f"{bound_type} bound")
            row += 1
        row += 1

    # Sample Data
    if report.sample_data is not None:
        ws.cell(row=row, column=1, value="SAMPLE DATA").font = section_font
        actual_n = sample_n if sample_n != -1 else report.sample_data.n_total
        ws.cell(row=row, column=3, value=f"(showing {min(actual_n, report.sample_data.n_sample)} of {report.sample_data.n_total})")
        row += 1

        # Headers - use X column names (not parameter names)
        sample_headers = ["Row"]
        if report.sample_data.x_column_names:
            sample_headers.extend(report.sample_data.x_column_names)
        else:
            sample_headers.extend([f"X{i+1}" for i in range(report.sample_data.X_sample.shape[1])])
        sample_headers.extend(["Y (Actual)", "Y (Pred)", "Residual"])

        for col, header in enumerate(sample_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        n_rows = min(actual_n, len(report.sample_data.y_sample))
        for i in range(n_rows):
            col = 1
            ws.cell(row=row, column=col, value=i+1).border = thin_border
            col += 1
            for j in range(report.sample_data.X_sample.shape[1]):
                ws.cell(row=row, column=col, value=f"{report.sample_data.X_sample[i, j]:.4f}").border = thin_border
                col += 1
            ws.cell(row=row, column=col, value=f"{report.sample_data.y_sample[i]:.4f}").border = thin_border
            col += 1
            ws.cell(row=row, column=col, value=f"{report.sample_data.y_pred_sample[i]:.4f}").border = thin_border
            col += 1
            residual = report.sample_data.y_sample[i] - report.sample_data.y_pred_sample[i]
            ws.cell(row=row, column=col, value=f"{residual:.4f}").border = thin_border
            row += 1
        row += 1

    # Alpha Trace Analysis
    if report.alpha_trace is not None:
        ws.cell(row=row, column=1, value="ALPHA TRACE ANALYSIS").font = section_font
        row += 1

        # Summary table header
        trace_headers = ["L1 Ratio", "Best Alpha", "Min Loss", "Converged", "Zero Coefs"]
        for col, header in enumerate(trace_headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        # Summary table rows
        for _, summary_row in report.alpha_trace.summary_df.iterrows():
            ws.cell(row=row, column=1, value=f"{summary_row['l1_ratio']:.1f}").border = thin_border
            ws.cell(row=row, column=2, value=f"{summary_row['best_alpha']:.6f}").border = thin_border
            ws.cell(row=row, column=3, value=f"{summary_row['min_loss']:.4f}").border = thin_border
            ws.cell(row=row, column=4, value=f"{summary_row['n_converged']}/{summary_row['n_total']}").border = thin_border
            n_zero = int(summary_row['n_zero_coefs']) if not np.isnan(summary_row['n_zero_coefs']) else "N/A"
            ws.cell(row=row, column=5, value=str(n_zero)).border = thin_border
            row += 1
        row += 1

        # Optimal hyperparameters
        opt = report.alpha_trace.optimal
        ws.cell(row=row, column=1, value="Optimal:").font = Font(bold=True)
        ws.cell(row=row, column=2, value=f"Alpha={opt['alpha']:.6f}")
        ws.cell(row=row, column=3, value=f"L1 Ratio={opt['l1_ratio']:.1f}")
        ws.cell(row=row, column=4, value=f"Loss={opt['loss_value']:.4f}")
        row += 2

    # Footer
    ws.cell(row=row, column=1, value=f"Report generated: {report.report_datetime}")
    ws.cell(row=row, column=4, value=f"CI Method: {report.ci_method}")

    # Adjust column widths
    for col in range(1, 10):
        ws.column_dimensions[chr(64 + col)].width = 15

    wb.save(filepath)


def _to_excel_multi_sheet(report: 'SummaryReport', filepath: str, sample_n: int):
    """Legacy multi-sheet Excel export."""
    import pandas as pd

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Model Equation sheet (first for prominence)
        if report.equation is not None:
            eq_data = {
                'Model Equation': [report.equation.text],
                'Is Custom Model': [report.equation.is_custom],
            }
            if report.equation.latex:
                eq_data['LaTeX'] = [report.equation.latex]
            if report.equation.source:
                eq_data['Source Code'] = [report.equation.source]
            eq_df = pd.DataFrame(eq_data)
            eq_df.to_excel(writer, sheet_name='Model Equation', index=False)

        # Coefficients sheet
        coef_df = _coefficients_to_dataframe(report.coefficients, report.intercept)
        coef_df.to_excel(writer, sheet_name='Coefficients', index=False)

        # Fit Statistics sheet
        fit_df = pd.DataFrame([report.fit_stats.to_dict()]).T
        fit_df.columns = ['Value']
        fit_df.index.name = 'Statistic'
        fit_df.to_excel(writer, sheet_name='Fit Statistics')

        # Model Specification sheet
        spec_dict = {
            'Model Type': report.model_spec.model_type,
            'Loss Function': report.model_spec.loss_function,
            'Alpha': report.model_spec.alpha,
            'L1 Ratio': report.model_spec.l1_ratio,
            'Method': report.model_spec.method,
            'Converged': report.model_spec.converged,
            'Final Objective': report.model_spec.final_objective,
            'Fit Date': str(report.model_spec.fit_datetime) if report.model_spec.fit_datetime else None,
            'Fit Duration (s)': report.model_spec.fit_duration_seconds,
        }
        spec_df = pd.DataFrame([spec_dict]).T
        spec_df.columns = ['Value']
        spec_df.index.name = 'Parameter'
        spec_df.to_excel(writer, sheet_name='Model Specification')

        # Data Summary sheet
        data_dict = {
            'N Samples': report.data_summary.n_samples,
            'N Features': report.data_summary.n_features,
            'Y Mean': report.data_summary.y_mean,
            'Y Std': report.data_summary.y_std,
            'Y Min': report.data_summary.y_min,
            'Y Max': report.data_summary.y_max,
        }
        data_df = pd.DataFrame([data_dict]).T
        data_df.columns = ['Value']
        data_df.index.name = 'Statistic'
        data_df.to_excel(writer, sheet_name='Data Summary')

        # Residual Analysis (if available)
        if report.residuals is not None:
            resid_dict = {
                'Mean': report.residuals.mean,
                'Std': report.residuals.std,
                'Min': report.residuals.min,
                'Max': report.residuals.max,
                'Skewness': report.residuals.skewness,
                'Kurtosis': report.residuals.kurtosis,
                'N Outliers': report.residuals.n_outliers,
            }
            resid_df = pd.DataFrame([resid_dict]).T
            resid_df.columns = ['Value']
            resid_df.index.name = 'Statistic'
            resid_df.to_excel(writer, sheet_name='Residual Analysis')

        # Sample Data sheet (if available)
        if report.sample_data is not None:
            sample_df = _sample_data_to_dataframe(report.sample_data, sample_n)
            sample_df.to_excel(writer, sheet_name='Sample Data', index=False)

        # Note about plots
        notes_df = pd.DataFrame({
            'Note': [
                'Diagnostic plots are available via report.plot_diagnostics()',
                'HTML export includes embedded plots: report.to_html("report.html")',
                f'Report generated: {report.report_datetime}',
                f'CI Method: {report.ci_method}'
            ]
        })
        notes_df.to_excel(writer, sheet_name='Notes', index=False)


def to_html(
    report: 'SummaryReport',
    filepath: Optional[str] = None,
    include_plots: bool = True,
    sample_n: int = 50,
    include_equation: bool = True,
    X: Optional[np.ndarray] = None,
    y: Optional[np.ndarray] = None,
) -> str:
    """
    Export report to HTML format with embedded plots and sample data.

    Parameters
    ----------
    report : SummaryReport
        The summary report to export
    filepath : str, optional
        If provided, write HTML to file. Otherwise return as string.
    include_plots : bool, default=True
        Embed diagnostic plots as base64 images
    sample_n : int, default=50
        Number of sample rows to include in sample data table. -1 for all (warns if >100)
    include_equation : bool, default=True
        Include model equation section
    X : np.ndarray, optional
        Full feature matrix for interactive data explorer. If None, uses sample_data.
    y : np.ndarray, optional
        Full target array for interactive data explorer. If None, uses sample_data.

    Returns
    -------
    str
        HTML content.
    """
    html_parts = []

    # CSS Styles with Plotly support
    html_parts.append("""
    <!DOCTYPE html>
    <html>
    <head>
        <meta charset="UTF-8">
        <title>Penalized-Constrained Regression Summary</title>
        <script src="https://cdn.plot.ly/plotly-2.27.0.min.js"></script>
        <script src="https://cdnjs.cloudflare.com/ajax/libs/mathjax/3.2.2/es5/tex-mml-chtml.min.js"></script>
        <style>
            body { font-family: 'Segoe UI', Arial, sans-serif; margin: 40px; background: #f5f5f5; }
            .container { max-width: 1200px; margin: 0 auto; background: white; padding: 30px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }
            h1 { color: #2c3e50; border-bottom: 3px solid #3498db; padding-bottom: 10px; }
            h2 { color: #34495e; margin-top: 30px; border-bottom: 1px solid #bdc3c7; padding-bottom: 5px; }
            h3 { color: #5d6d7e; margin-top: 20px; }
            table { border-collapse: collapse; width: 100%; margin: 15px 0; }
            th, td { border: 1px solid #ddd; padding: 10px; text-align: left; }
            th { background-color: #3498db; color: white; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            tr:hover { background-color: #f5f5f5; }
            .stat-value { font-family: 'Consolas', monospace; }
            .at-bound { color: #e74c3c; font-weight: bold; }
            .free { color: #27ae60; }
            .converged-true { color: #27ae60; }
            .converged-false { color: #e74c3c; }
            .metric-card { display: inline-block; background: #ecf0f1; padding: 15px 25px; margin: 5px; border-radius: 5px; text-align: center; }
            .metric-value { font-size: 1.5em; font-weight: bold; color: #2c3e50; }
            .metric-label { font-size: 0.9em; color: #7f8c8d; }
            .equation-box { background: #f8f9fa; border: 1px solid #e9ecef; border-radius: 5px; padding: 15px; margin: 15px 0; font-family: 'Consolas', monospace; }
            .equation-latex { font-style: italic; color: #6c757d; margin-top: 10px; }
            .source-code { background: #2d2d2d; color: #f8f8f2; padding: 15px; border-radius: 5px; overflow-x: auto; white-space: pre-wrap; font-size: 0.9em; }
            .plot-container { text-align: center; margin: 20px 0; }
            .plot-container img { max-width: 100%; border: 1px solid #ddd; border-radius: 5px; }
            .plotly-container { width: 100%; margin: 20px 0; }
            .sample-data { max-height: 400px; overflow-y: auto; }
            .warning { background: #fff3cd; border: 1px solid #ffc107; color: #856404; padding: 10px; border-radius: 5px; margin: 10px 0; }
            .info-box { background: #d1ecf1; border: 1px solid #bee5eb; color: #0c5460; padding: 10px; border-radius: 5px; margin: 10px 0; }
            .controls { background: #f8f9fa; padding: 15px; border-radius: 5px; margin: 15px 0; display: flex; flex-wrap: wrap; gap: 15px; align-items: center; }
            .control-group { display: flex; flex-direction: column; }
            .control-group label { font-weight: bold; margin-bottom: 5px; font-size: 0.9em; color: #34495e; }
            .control-group select { padding: 8px 12px; border: 1px solid #ddd; border-radius: 4px; background: white; min-width: 150px; }
            .bootstrap-section { background: #f8f9fa; padding: 20px; border-radius: 8px; margin: 20px 0; }
        </style>
    </head>
    <body>
    <div class="container">
    """)

    # Title
    html_parts.append("<h1>Penalized-Constrained Regression Summary</h1>")

    # Key Metrics Cards
    html_parts.append('<div style="text-align: center; margin: 20px 0;">')
    html_parts.append(f'<div class="metric-card"><div class="metric-value">{report.fit_stats.r2:.4f}</div><div class="metric-label">R2</div></div>')
    html_parts.append(f'<div class="metric-card"><div class="metric-value">{report.fit_stats.spe:.2%}</div><div class="metric-label">SPE</div></div>')
    html_parts.append(f'<div class="metric-card"><div class="metric-value">{report.fit_stats.mape:.2%}</div><div class="metric-label">MAPE</div></div>')
    html_parts.append(f'<div class="metric-card"><div class="metric-value">{report.constraints.n_active}</div><div class="metric-label">Active Constraints</div></div>')
    html_parts.append('</div>')

    # Model Equation (NEW)
    if include_equation and report.equation is not None:
        html_parts.append("<h2>Model Equation</h2>")
        html_parts.append(f'<div class="equation-box">')
        # Use <pre> style to preserve newlines and spacing in equation text
        escaped_text = _escape_html(report.equation.text)
        html_parts.append(f'<pre style="margin: 0; font-family: Consolas, monospace; white-space: pre-wrap;">{escaped_text}</pre>')
        if report.equation.latex:
            # Use MathJax display mode delimiters for rendered equation
            html_parts.append(f'<div class="equation-latex" style="margin-top: 15px; font-size: 1.2em;">\\[{report.equation.latex}\\]</div>')
        html_parts.append('</div>')

        if report.equation.source:
            html_parts.append("<h3>Source Code</h3>")
            html_parts.append(f'<pre class="source-code">{_escape_html(report.equation.source)}</pre>')

    # Model Specification - All input parameters for reproducibility
    html_parts.append("<h2>Model Specification</h2>")
    html_parts.append("<h3>Constructor Parameters</h3>")
    html_parts.append("<table>")
    html_parts.append(f"<tr><td>Model Type</td><td class='stat-value'>{report.model_spec.model_type}</td></tr>")
    html_parts.append(f"<tr><td>Loss Function</td><td class='stat-value'>{report.model_spec.loss_function}</td></tr>")
    html_parts.append(f"<tr><td>Alpha</td><td class='stat-value'>{report.model_spec.alpha}</td></tr>")
    html_parts.append(f"<tr><td>L1 Ratio</td><td class='stat-value'>{report.model_spec.l1_ratio}</td></tr>")
    html_parts.append(f"<tr><td>Fit Intercept</td><td class='stat-value'>{report.model_spec.fit_intercept}</td></tr>")
    html_parts.append(f"<tr><td>Scale</td><td class='stat-value'>{report.model_spec.scale}</td></tr>")
    # Bounds
    if report.model_spec.bounds:
        bounds_str = str(report.model_spec.bounds)
        html_parts.append(f"<tr><td>Bounds</td><td class='stat-value'>{_escape_html(bounds_str)}</td></tr>")
    if report.model_spec.intercept_bounds:
        html_parts.append(f"<tr><td>Intercept Bounds</td><td class='stat-value'>{report.model_spec.intercept_bounds}</td></tr>")
    # Naming
    if report.model_spec.coef_names:
        html_parts.append(f"<tr><td>Coefficient Names</td><td class='stat-value'>{report.model_spec.coef_names}</td></tr>")
    if report.model_spec.penalty_exclude:
        html_parts.append(f"<tr><td>Penalty Exclude</td><td class='stat-value'>{report.model_spec.penalty_exclude}</td></tr>")
    # Optimization
    html_parts.append(f"<tr><td>Method</td><td class='stat-value'>{report.model_spec.method}</td></tr>")
    html_parts.append(f"<tr><td>Max Iterations</td><td class='stat-value'>{report.model_spec.max_iter}</td></tr>")
    html_parts.append(f"<tr><td>Tolerance</td><td class='stat-value'>{report.model_spec.tol}</td></tr>")
    # x0 initial values
    x0_str = str(report.model_spec.x0) if report.model_spec.x0 else 'ols'
    html_parts.append(f"<tr><td>x0 (Initial Values)</td><td class='stat-value'>{_escape_html(x0_str)}</td></tr>")
    # Custom functions
    if report.model_spec.has_custom_prediction_fn:
        html_parts.append(f"<tr><td>Custom Prediction Function</td><td class='stat-value'>Yes</td></tr>")
    if report.model_spec.has_custom_loss_fn:
        html_parts.append(f"<tr><td>Custom Loss Function</td><td class='stat-value'>Yes</td></tr>")
    html_parts.append("</table>")

    # Fit Results
    html_parts.append("<h3>Fit Results</h3>")
    html_parts.append("<table>")
    converged_class = 'converged-true' if report.model_spec.converged else 'converged-false'
    html_parts.append(f"<tr><td>Converged</td><td class='stat-value {converged_class}'>{report.model_spec.converged}</td></tr>")
    if report.model_spec.n_iterations:
        html_parts.append(f"<tr><td>Iterations</td><td class='stat-value'>{report.model_spec.n_iterations}</td></tr>")
    if report.model_spec.final_objective:
        html_parts.append(f"<tr><td>Final Objective</td><td class='stat-value'>{report.model_spec.final_objective:.6f}</td></tr>")
    if report.model_spec.fit_datetime:
        html_parts.append(f"<tr><td>Fit Date</td><td class='stat-value'>{report.model_spec.fit_datetime}</td></tr>")
    if report.model_spec.fit_duration_seconds:
        html_parts.append(f"<tr><td>Fit Duration</td><td class='stat-value'>{report.model_spec.fit_duration_seconds:.2f}s</td></tr>")
    html_parts.append("</table>")

    # Data Summary
    html_parts.append("<h2>Data Summary</h2>")
    html_parts.append("<table>")
    html_parts.append(f"<tr><td>N Samples</td><td class='stat-value'>{report.data_summary.n_samples}</td></tr>")
    html_parts.append(f"<tr><td>N Features</td><td class='stat-value'>{report.data_summary.n_features}</td></tr>")
    html_parts.append(f"<tr><td>Y Mean</td><td class='stat-value'>{report.data_summary.y_mean:.4f}</td></tr>")
    html_parts.append(f"<tr><td>Y Std</td><td class='stat-value'>{report.data_summary.y_std:.4f}</td></tr>")
    html_parts.append(f"<tr><td>Y Range</td><td class='stat-value'>[{report.data_summary.y_min:.4f}, {report.data_summary.y_max:.4f}]</td></tr>")
    html_parts.append("</table>")

    # Coefficients - show both Hessian and Bootstrap SE/CI
    html_parts.append("<h2>Coefficients</h2>")

    # Determine what columns we have
    has_hessian = any(c.hessian_se is not None for c in report.coefficients)
    has_bootstrap = any(c.bootstrap_se is not None for c in report.coefficients)

    html_parts.append("<table>")

    # Build header based on available data
    header_parts = ["<tr><th>Parameter</th><th>Value</th>"]
    if has_hessian:
        header_parts.append("<th>Hessian SE</th><th>Hessian 95% CI</th>")
    if has_bootstrap:
        header_parts.append("<th>Bootstrap SE</th><th>Bootstrap 95% CI</th>")
    header_parts.append("<th>Bounds</th><th>Status</th></tr>")
    html_parts.append("".join(header_parts))

    all_coefs = list(report.coefficients)
    if report.intercept is not None:
        all_coefs.append(report.intercept)

    for c in all_coefs:
        lb = f"{c.lower_bound:.4f}" if np.isfinite(c.lower_bound) else "-inf"
        ub = f"{c.upper_bound:.4f}" if np.isfinite(c.upper_bound) else "+inf"
        bounds_str = f"[{lb}, {ub}]"
        status_class = 'at-bound' if c.is_constrained else 'free'

        row_parts = [f"<tr><td>{c.name}</td><td class='stat-value'>{c.value:.6f}</td>"]

        # Hessian columns
        if has_hessian:
            if c.hessian_se is not None:
                hess_ci_str = f"[{c.hessian_ci_lower:.4f}, {c.hessian_ci_upper:.4f}]"
                row_parts.append(f"<td class='stat-value'>{c.hessian_se:.4f}</td><td class='stat-value'>{hess_ci_str}</td>")
            else:
                row_parts.append("<td>-</td><td>-</td>")

        # Bootstrap columns
        if has_bootstrap:
            if c.bootstrap_se is not None:
                boot_ci_str = f"[{c.bootstrap_ci_lower:.4f}, {c.bootstrap_ci_upper:.4f}]"
                row_parts.append(f"<td class='stat-value'>{c.bootstrap_se:.4f}</td><td class='stat-value'>{boot_ci_str}</td>")
            else:
                row_parts.append("<td>-</td><td>-</td>")

        row_parts.append(f"<td>{bounds_str}</td><td class='{status_class}'>{c.bound_status}</td></tr>")
        html_parts.append("".join(row_parts))

    html_parts.append("</table>")

    # Fit Statistics
    html_parts.append("<h2>Fit Statistics</h2>")
    html_parts.append("<table>")
    html_parts.append(f"<tr><td>R2</td><td class='stat-value'>{report.fit_stats.r2:.4f}</td></tr>")
    html_parts.append(f"<tr><td>Adjusted R2</td><td class='stat-value'>{report.fit_stats.adj_r2:.4f}</td></tr>")
    html_parts.append(f"<tr><td>SEE</td><td class='stat-value'>{report.fit_stats.see:.4f}</td></tr>")
    html_parts.append(f"<tr><td>SPE</td><td class='stat-value'>{report.fit_stats.spe:.2%}</td></tr>")
    html_parts.append(f"<tr><td>MAPE</td><td class='stat-value'>{report.fit_stats.mape:.2%}</td></tr>")
    html_parts.append(f"<tr><td>RMSE</td><td class='stat-value'>{report.fit_stats.rmse:.4f}</td></tr>")
    if report.fit_stats.mse is not None:
        html_parts.append(f"<tr><td>MSE</td><td class='stat-value'>{report.fit_stats.mse:.4f}</td></tr>")
    if report.fit_stats.mae is not None:
        html_parts.append(f"<tr><td>MAE</td><td class='stat-value'>{report.fit_stats.mae:.4f}</td></tr>")
    html_parts.append(f"<tr><td>CV</td><td class='stat-value'>{report.fit_stats.cv:.2%}</td></tr>")
    html_parts.append(f"<tr><td>GDF ({report.fit_stats.gdf_method})</td><td class='stat-value'>{report.fit_stats.gdf:.1f}</td></tr>")
    if report.fit_stats.aic is not None:
        html_parts.append(f"<tr><td>AIC</td><td class='stat-value'>{report.fit_stats.aic:.4f}</td></tr>")
    if report.fit_stats.bic is not None:
        html_parts.append(f"<tr><td>BIC</td><td class='stat-value'>{report.fit_stats.bic:.4f}</td></tr>")
    if report.fit_stats.f_statistic is not None:
        pval_str = f"{report.fit_stats.f_pvalue:.4e}" if report.fit_stats.f_pvalue else "N/A"
        html_parts.append(f"<tr><td>F-statistic</td><td class='stat-value'>{report.fit_stats.f_statistic:.4f} (p={pval_str})</td></tr>")
    if report.fit_stats.durbin_watson is not None:
        html_parts.append(f"<tr><td>Durbin-Watson</td><td class='stat-value'>{report.fit_stats.durbin_watson:.4f}</td></tr>")
    html_parts.append("</table>")

    # Residual Analysis
    if report.residuals is not None:
        html_parts.append("<h2>Residual Analysis</h2>")
        html_parts.append("<table>")
        html_parts.append(f"<tr><td>Mean</td><td class='stat-value'>{report.residuals.mean:.4f}</td></tr>")
        html_parts.append(f"<tr><td>Std</td><td class='stat-value'>{report.residuals.std:.4f}</td></tr>")
        html_parts.append(f"<tr><td>Range</td><td class='stat-value'>[{report.residuals.min:.4f}, {report.residuals.max:.4f}]</td></tr>")
        html_parts.append(f"<tr><td>Skewness</td><td class='stat-value'>{report.residuals.skewness:.4f}</td></tr>")
        html_parts.append(f"<tr><td>Kurtosis</td><td class='stat-value'>{report.residuals.kurtosis:.4f}</td></tr>")
        html_parts.append(f"<tr><td>Outliers (|r| > 2 sigma)</td><td class='stat-value'>{report.residuals.n_outliers}</td></tr>")
        html_parts.append("</table>")

    # Active Constraints
    if report.constraints.active_constraints:
        html_parts.append("<h2>Active Constraints</h2>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Bound Type</th></tr>")
        for name, bound_type in report.constraints.active_constraints:
            html_parts.append(f"<tr><td>{name}</td><td>{bound_type}</td></tr>")
        html_parts.append("</table>")

    # Sample Data (NEW)
    if report.sample_data is not None:
        actual_n = sample_n if sample_n != -1 else report.sample_data.n_total
        if sample_n == -1 and report.sample_data.n_total > 100:
            html_parts.append('<div class="warning">Warning: Large dataset ({} rows). Consider using sample_n parameter to limit display.</div>'.format(report.sample_data.n_total))

        html_parts.append("<h2>Sample Data</h2>")
        html_parts.append(f"<p>Showing {min(actual_n, report.sample_data.n_sample)} of {report.sample_data.n_total} observations</p>")
        html_parts.append('<div class="sample-data">')
        html_parts.append("<table>")

        # Header - use X column names (not parameter names)
        header_parts = ["<tr><th>Row</th>"]
        if report.sample_data.x_column_names:
            for name in report.sample_data.x_column_names:
                header_parts.append(f"<th>{name}</th>")
        else:
            for i in range(report.sample_data.X_sample.shape[1]):
                header_parts.append(f"<th>X{i+1}</th>")
        header_parts.append("<th>Y (Actual)</th><th>Y (Predicted)</th><th>Residual</th></tr>")
        html_parts.append("".join(header_parts))

        # Data rows
        n_rows = min(actual_n, len(report.sample_data.y_sample))
        for i in range(n_rows):
            row_parts = [f"<tr><td>{i+1}</td>"]
            for j in range(report.sample_data.X_sample.shape[1]):
                row_parts.append(f"<td class='stat-value'>{report.sample_data.X_sample[i, j]:.4f}</td>")
            y_actual = report.sample_data.y_sample[i]
            y_pred = report.sample_data.y_pred_sample[i]
            residual = y_actual - y_pred
            row_parts.append(f"<td class='stat-value'>{y_actual:.4f}</td>")
            row_parts.append(f"<td class='stat-value'>{y_pred:.4f}</td>")
            row_parts.append(f"<td class='stat-value'>{residual:.4f}</td></tr>")
            html_parts.append("".join(row_parts))

        html_parts.append("</table>")
        html_parts.append("</div>")

    # Diagnostic Plots
    if include_plots and report.residuals is not None and report.residuals.residuals is not None:
        from .plotting import generate_embedded_plots

        html_parts.append("<h2>Diagnostic Plots</h2>")
        plot_data = generate_embedded_plots(report.residuals, report.coefficients)

        if plot_data.get('diagnostics'):
            html_parts.append('<div class="plot-container">')
            html_parts.append(f'<img src="data:image/{plot_data["format"]};base64,{plot_data["diagnostics"]}" alt="Diagnostic Plots">')
            html_parts.append('</div>')
        elif plot_data.get('error'):
            html_parts.append(f'<p class="warning">Could not generate plots: {plot_data["error"]}</p>')

    # Alpha Trace Analysis
    if report.alpha_trace is not None:
        html_parts.append("<h2>Alpha Trace Analysis</h2>")
        html_parts.append("<p>Coefficient paths across different regularization strengths (alpha) and L1 ratios.</p>")

        # Summary table
        html_parts.append("<h3>Summary by L1 Ratio</h3>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>L1 Ratio</th><th>Best Alpha</th><th>Min Loss</th><th>Converged</th><th>Zero Coefs</th></tr>")
        for _, row in report.alpha_trace.summary_df.iterrows():
            html_parts.append(f"<tr><td>{row['l1_ratio']:.1f}</td><td class='stat-value'>{row['best_alpha']:.6f}</td>"
                            f"<td class='stat-value'>{row['min_loss']:.4f}</td>"
                            f"<td>{row['n_converged']}/{row['n_total']}</td>"
                            f"<td>{int(row['n_zero_coefs']) if not np.isnan(row['n_zero_coefs']) else 'N/A'}</td></tr>")
        html_parts.append("</table>")

        # Optimal parameters
        opt = report.alpha_trace.optimal
        html_parts.append("<h3>Optimal Hyperparameters</h3>")
        html_parts.append(f"<p><strong>Alpha:</strong> {opt['alpha']:.6f} | <strong>L1 Ratio:</strong> {opt['l1_ratio']:.1f} | <strong>Loss:</strong> {opt['loss_value']:.4f}</p>")

        # Embed alpha trace plot if available
        if include_plots:
            try:
                from .alpha_trace import plot_alpha_trace
                from .plotting import figure_to_base64
                import matplotlib
                matplotlib.use('Agg')

                fig = plot_alpha_trace(report.alpha_trace.trace_df, show_legend=True)
                plot_b64 = figure_to_base64(fig)
                import matplotlib.pyplot as plt
                plt.close(fig)

                html_parts.append('<div class="plot-container">')
                html_parts.append(f'<img src="data:image/png;base64,{plot_b64}" alt="Alpha Trace Plot">')
                html_parts.append('</div>')
            except Exception as e:
                html_parts.append(f'<p class="warning">Could not generate alpha trace plot: {e}</p>')

    # Bootstrap Analysis Section
    html_parts.append(_generate_bootstrap_section_html(report))

    # Interactive Data Explorer (X vs Y scatter plot with Plotly)
    # Use full X, y if provided, otherwise fall back to sample_data
    if X is not None and y is not None:
        html_parts.append(_generate_interactive_scatter_html(report, sample_n, X, y))
    elif report.sample_data is not None:
        html_parts.append(_generate_interactive_scatter_html(report, sample_n, None, None))

    # Footer
    html_parts.append(f"<p style='color: #7f8c8d; font-size: 0.9em; margin-top: 30px;'>Report generated: {report.report_datetime} | CI Method: {report.ci_method}</p>")
    html_parts.append("</div></body></html>")

    html_content = "\n".join(html_parts)

    if filepath is not None:
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)

    return html_content


def to_pdf(report: 'SummaryReport', filepath: str, **kwargs):
    """
    Export report to PDF format.

    Parameters
    ----------
    report : SummaryReport
        The summary report to export
    filepath : str
        Path to output PDF file.
    **kwargs
        Additional arguments passed to to_html()

    Notes
    -----
    Requires weasyprint or pdfkit to be installed.
    Falls back to HTML if PDF generation fails.
    """
    html_content = to_html(report, **kwargs)

    # Try weasyprint first
    try:
        from weasyprint import HTML
        HTML(string=html_content).write_pdf(filepath)
        return
    except ImportError:
        pass

    # Try pdfkit
    try:
        import pdfkit
        pdfkit.from_string(html_content, filepath)
        return
    except ImportError:
        pass

    # Fall back to saving HTML with PDF extension warning
    warnings.warn(
        "Neither weasyprint nor pdfkit is installed. "
        "Install with: pip install weasyprint or pip install pdfkit. "
        "Saving as HTML instead.",
        UserWarning
    )
    html_filepath = filepath.replace('.pdf', '.html')
    to_html(report, html_filepath, **kwargs)


def _coefficients_to_dataframe(coefficients: List['CoefficientInfo'], intercept: Optional['CoefficientInfo']):
    """Convert coefficients to pandas DataFrame."""
    import pandas as pd

    rows = []
    for c in coefficients:
        row = {
            'Name': c.name,
            'Value': c.value,
            'Lower Bound': c.lower_bound,
            'Upper Bound': c.upper_bound,
            'Status': c.bound_status,
        }
        if c.se is not None:
            row['SE'] = c.se
        if c.ci_lower is not None:
            row['CI Lower'] = c.ci_lower
            row['CI Upper'] = c.ci_upper
        rows.append(row)

    if intercept is not None:
        row = {
            'Name': 'Intercept',
            'Value': intercept.value,
            'Lower Bound': intercept.lower_bound,
            'Upper Bound': intercept.upper_bound,
            'Status': intercept.bound_status,
        }
        if intercept.se is not None:
            row['SE'] = intercept.se
        if intercept.ci_lower is not None:
            row['CI Lower'] = intercept.ci_lower
            row['CI Upper'] = intercept.ci_upper
        rows.append(row)

    return pd.DataFrame(rows)


def _sample_data_to_dataframe(sample_data: 'SampleData', sample_n: int):
    """Convert sample data to pandas DataFrame."""
    import pandas as pd

    n_rows = sample_n if sample_n != -1 else sample_data.n_sample
    n_rows = min(n_rows, sample_data.n_sample)

    data = {}

    # Features - use X column names (not parameter names)
    if sample_data.x_column_names:
        for i, name in enumerate(sample_data.x_column_names):
            data[name] = sample_data.X_sample[:n_rows, i]
    else:
        for i in range(sample_data.X_sample.shape[1]):
            data[f'X{i+1}'] = sample_data.X_sample[:n_rows, i]

    # Target and predictions
    data['Y_Actual'] = sample_data.y_sample[:n_rows]
    data['Y_Predicted'] = sample_data.y_pred_sample[:n_rows]
    data['Residual'] = sample_data.y_sample[:n_rows] - sample_data.y_pred_sample[:n_rows]

    return pd.DataFrame(data)


def _escape_html(text: str) -> str:
    """Escape HTML special characters."""
    return (text
            .replace('&', '&amp;')
            .replace('<', '&lt;')
            .replace('>', '&gt;')
            .replace('"', '&quot;')
            .replace("'", '&#39;'))


def _generate_bootstrap_section_html(report: 'SummaryReport') -> str:
    """
    Generate HTML for the Bootstrap Analysis section.

    Parameters
    ----------
    report : SummaryReport
        The summary report containing bootstrap_results

    Returns
    -------
    str
        HTML string for the bootstrap section (empty if no bootstrap results)
    """
    if report.bootstrap_results is None:
        # Return empty placeholder section
        return """
        <h2>Bootstrap Analysis</h2>
        <div class="bootstrap-section">
            <p class="info-box">Bootstrap analysis was not performed. To include bootstrap confidence intervals,
            use <code>bootstrap=True</code> when generating the report.</p>
        </div>
        """

    bootstrap = report.bootstrap_results
    html_parts = []

    html_parts.append("<h2>Bootstrap Analysis</h2>")
    html_parts.append('<div class="bootstrap-section">')

    # Determine what we have
    has_constrained = bootstrap.constrained is not None
    has_unconstrained = bootstrap.unconstrained is not None

    # Summary info - use whichever is available for n_successful
    n_successful = bootstrap.constrained.n_successful if has_constrained else bootstrap.unconstrained.n_successful
    html_parts.append(f"""
    <p><strong>Bootstrap samples:</strong> {n_successful} of {bootstrap.n_bootstrap} successful |
    <strong>Confidence level:</strong> {bootstrap.confidence:.0%}</p>
    """)

    # Determine number of coefficients from whichever result is available
    if has_constrained:
        n_coefs = len(bootstrap.constrained.coef_mean)
    else:
        n_coefs = len(bootstrap.unconstrained.coef_mean)

    names = bootstrap.feature_names if bootstrap.feature_names else [f'coef_{i}' for i in range(n_coefs)]

    if has_constrained and has_unconstrained:
        # Model IS constrained: show both tables and comparison
        html_parts.append("""
        <p class="info-box">Bootstrap was run in two modes: <strong>Constrained</strong> (with bounds and regularization)
        and <strong>Unconstrained</strong> (no bounds, alpha=0) to show the effect of constraints on uncertainty estimates.</p>
        """)

        # Constrained Bootstrap Statistics
        constrained = bootstrap.constrained
        html_parts.append("<h3>Constrained Bootstrap (with bounds and alpha)</h3>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Mean</th><th>Std</th><th>CI Lower</th><th>CI Upper</th><th>CI Width</th></tr>")

        for i, name in enumerate(names):
            ci_width = constrained.coef_ci_upper[i] - constrained.coef_ci_lower[i]
            html_parts.append(f"""
            <tr>
                <td>{name}</td>
                <td class='stat-value'>{constrained.coef_mean[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_std[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_ci_lower[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_ci_upper[i]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)

        if constrained.intercept_mean is not None and constrained.intercept_ci is not None:
            ci_width = constrained.intercept_ci[1] - constrained.intercept_ci[0]
            html_parts.append(f"""
            <tr>
                <td>Intercept</td>
                <td class='stat-value'>{constrained.intercept_mean:.6f}</td>
                <td class='stat-value'>{constrained.intercept_std:.6f}</td>
                <td class='stat-value'>{constrained.intercept_ci[0]:.6f}</td>
                <td class='stat-value'>{constrained.intercept_ci[1]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)
        html_parts.append("</table>")

        # Unconstrained Bootstrap Statistics
        unconstrained = bootstrap.unconstrained
        html_parts.append("<h3>Unconstrained Bootstrap (no bounds, alpha=0)</h3>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Mean</th><th>Std</th><th>CI Lower</th><th>CI Upper</th><th>CI Width</th></tr>")

        for i, name in enumerate(names):
            ci_width = unconstrained.coef_ci_upper[i] - unconstrained.coef_ci_lower[i]
            html_parts.append(f"""
            <tr>
                <td>{name}</td>
                <td class='stat-value'>{unconstrained.coef_mean[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_std[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_ci_lower[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_ci_upper[i]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)

        if unconstrained.intercept_mean is not None and unconstrained.intercept_ci is not None:
            ci_width = unconstrained.intercept_ci[1] - unconstrained.intercept_ci[0]
            html_parts.append(f"""
            <tr>
                <td>Intercept</td>
                <td class='stat-value'>{unconstrained.intercept_mean:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_std:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_ci[0]:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_ci[1]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)
        html_parts.append("</table>")

        # Comparison table
        html_parts.append("<h3>Constraint Effect on Uncertainty</h3>")
        html_parts.append("<p>Comparison of CI widths between constrained and unconstrained bootstrap:</p>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Constrained CI Width</th><th>Unconstrained CI Width</th><th>Reduction</th></tr>")

        for i, name in enumerate(names):
            const_width = constrained.coef_ci_upper[i] - constrained.coef_ci_lower[i]
            unconst_width = unconstrained.coef_ci_upper[i] - unconstrained.coef_ci_lower[i]
            reduction = ((unconst_width - const_width) / unconst_width * 100) if unconst_width > 0 else 0
            html_parts.append(f"""
            <tr>
                <td>{name}</td>
                <td class='stat-value'>{const_width:.6f}</td>
                <td class='stat-value'>{unconst_width:.6f}</td>
                <td class='stat-value'>{reduction:+.1f}%</td>
            </tr>
            """)
        html_parts.append("</table>")

    elif has_unconstrained and not has_constrained:
        # Model has NO constraints: only show unconstrained results
        html_parts.append("""
        <p class="info-box">Model has no constraints (no bounds and alpha=0), so only unconstrained bootstrap was performed.
        Constrained bootstrap comparison is not applicable.</p>
        """)

        unconstrained = bootstrap.unconstrained
        html_parts.append("<h3>Bootstrap Results (Unconstrained Model)</h3>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Mean</th><th>Std</th><th>CI Lower</th><th>CI Upper</th><th>CI Width</th></tr>")

        for i, name in enumerate(names):
            ci_width = unconstrained.coef_ci_upper[i] - unconstrained.coef_ci_lower[i]
            html_parts.append(f"""
            <tr>
                <td>{name}</td>
                <td class='stat-value'>{unconstrained.coef_mean[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_std[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_ci_lower[i]:.6f}</td>
                <td class='stat-value'>{unconstrained.coef_ci_upper[i]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)

        if unconstrained.intercept_mean is not None and unconstrained.intercept_ci is not None:
            ci_width = unconstrained.intercept_ci[1] - unconstrained.intercept_ci[0]
            html_parts.append(f"""
            <tr>
                <td>Intercept</td>
                <td class='stat-value'>{unconstrained.intercept_mean:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_std:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_ci[0]:.6f}</td>
                <td class='stat-value'>{unconstrained.intercept_ci[1]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)
        html_parts.append("</table>")

    elif has_constrained and not has_unconstrained:
        # Model IS constrained but unconstrained bootstrap failed
        # (e.g., model requires bounds to converge, custom prediction_fn, etc.)
        html_parts.append("""
        <p class="info-box">Model has constraints. Unconstrained bootstrap comparison was attempted but failed
        (model may require constraints to converge properly). Showing constrained bootstrap results only.</p>
        """)

        constrained = bootstrap.constrained
        html_parts.append("<h3>Constrained Bootstrap (with bounds and alpha)</h3>")
        html_parts.append("<table>")
        html_parts.append("<tr><th>Parameter</th><th>Mean</th><th>Std</th><th>CI Lower</th><th>CI Upper</th><th>CI Width</th></tr>")

        for i, name in enumerate(names):
            ci_width = constrained.coef_ci_upper[i] - constrained.coef_ci_lower[i]
            html_parts.append(f"""
            <tr>
                <td>{name}</td>
                <td class='stat-value'>{constrained.coef_mean[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_std[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_ci_lower[i]:.6f}</td>
                <td class='stat-value'>{constrained.coef_ci_upper[i]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)

        if constrained.intercept_mean is not None and constrained.intercept_ci is not None:
            ci_width = constrained.intercept_ci[1] - constrained.intercept_ci[0]
            html_parts.append(f"""
            <tr>
                <td>Intercept</td>
                <td class='stat-value'>{constrained.intercept_mean:.6f}</td>
                <td class='stat-value'>{constrained.intercept_std:.6f}</td>
                <td class='stat-value'>{constrained.intercept_ci[0]:.6f}</td>
                <td class='stat-value'>{constrained.intercept_ci[1]:.6f}</td>
                <td class='stat-value'>{ci_width:.6f}</td>
            </tr>
            """)
        html_parts.append("</table>")

    else:
        # No results available
        html_parts.append("<p class='warning'>Bootstrap analysis failed to produce results.</p>")

    # Add bootstrap distribution plots using Plotly
    html_parts.append("<h3>Bootstrap Coefficient Distributions</h3>")
    html_parts.append('<div id="bootstrap-distributions" class="plotly-container"></div>')

    # Generate Plotly JavaScript for bootstrap histograms
    import json

    n_cols = min(3, n_coefs)
    n_rows = (n_coefs + n_cols - 1) // n_cols

    traces = []
    for i in range(n_coefs):
        name = names[i] if i < len(names) else f'coef_{i}'

        # Constrained distribution (if available)
        if has_constrained:
            coef_samples = bootstrap.constrained.bootstrap_coefs[:, i].tolist()
            traces.append({
                'type': 'histogram',
                'x': coef_samples,
                'name': f'{name} (constrained)',
                'opacity': 0.7,
                'marker': {'color': '#3498db'},
                'xaxis': f'x{i+1}' if i > 0 else 'x',
                'yaxis': f'y{i+1}' if i > 0 else 'y',
                'legendgroup': name,
            })

        # Unconstrained distribution (if available)
        if has_unconstrained:
            unconst_samples = bootstrap.unconstrained.bootstrap_coefs[:, i].tolist()
            # Use blue if no constrained, red if showing both
            color = '#e74c3c' if has_constrained else '#3498db'
            opacity = 0.5 if has_constrained else 0.7
            label = f'{name} (unconstrained)' if has_constrained else name
            traces.append({
                'type': 'histogram',
                'x': unconst_samples,
                'name': label,
                'opacity': opacity,
                'marker': {'color': color},
                'xaxis': f'x{i+1}' if i > 0 else 'x',
                'yaxis': f'y{i+1}' if i > 0 else 'y',
                'legendgroup': name,
            })

    # Create layout with subplots
    if has_constrained and has_unconstrained:
        title = 'Bootstrap Coefficient Distributions (blue=constrained, red=unconstrained)'
    else:
        title = 'Bootstrap Coefficient Distributions'

    layout = {
        'title': title,
        'showlegend': False,
        'height': 200 * n_rows + 100,
        'barmode': 'overlay',
    }

    # Create grid layout
    for i in range(n_coefs):
        row = i // n_cols
        col = i % n_cols
        x_domain = [col / n_cols + 0.02, (col + 1) / n_cols - 0.02]
        y_domain = [1 - (row + 1) / n_rows + 0.05, 1 - row / n_rows - 0.05]

        axis_suffix = str(i + 1) if i > 0 else ''
        name = names[i] if i < len(names) else f'coef_{i}'
        layout[f'xaxis{axis_suffix}'] = {
            'domain': x_domain,
            'title': name,
            'anchor': f'y{axis_suffix}' if axis_suffix else 'y',
        }
        layout[f'yaxis{axis_suffix}'] = {
            'domain': y_domain,
            'anchor': f'x{axis_suffix}' if axis_suffix else 'x',
        }

    html_parts.append(f"""
    <script>
        var bootstrapTraces = {json.dumps(traces)};
        var bootstrapLayout = {json.dumps(layout)};
        Plotly.newPlot('bootstrap-distributions', bootstrapTraces, bootstrapLayout, {{responsive: true}});
    </script>
    """)

    html_parts.append('</div>')
    return "\n".join(html_parts)


def _generate_interactive_scatter_html(
    report: 'SummaryReport',
    sample_n: int,
    X: Optional[np.ndarray] = None,
    y: Optional[np.ndarray] = None
) -> str:
    """
    Generate HTML for interactive X vs Y scatter plot with variable selectors.

    Parameters
    ----------
    report : SummaryReport
        The summary report containing sample_data
    sample_n : int
        Number of sample rows to include (only used for sample_data fallback display)
    X : np.ndarray, optional
        Full feature matrix. If provided, uses this instead of sample_data for explorer.
    y : np.ndarray, optional
        Full target array. If provided, uses this instead of sample_data for explorer.

    Returns
    -------
    str
        HTML string for the interactive scatter plot section
    """
    import json

    # Determine data source
    if X is not None and y is not None:
        # Use full dataset
        X_arr = np.asarray(X)
        y_arr = np.asarray(y).ravel()

        # Get X column names (not parameter names) - these label the columns in the X matrix
        if hasattr(X, 'columns'):
            x_col_names = list(X.columns)
        elif report.sample_data and report.sample_data.x_column_names:
            # Try to use X column names from sample_data if they match
            if len(report.sample_data.x_column_names) == X_arr.shape[1]:
                x_col_names = report.sample_data.x_column_names
            else:
                x_col_names = [f'X{i+1}' for i in range(X_arr.shape[1])]
        else:
            x_col_names = [f'X{i+1}' for i in range(X_arr.shape[1])]

        actual_n = len(y_arr)

        # Get predictions if residuals are available
        if report.residuals is not None and report.residuals.y_pred is not None:
            y_pred = report.residuals.y_pred
        else:
            y_pred = np.zeros_like(y_arr)  # Fallback

        data_source = "full dataset"
    elif report.sample_data is not None:
        # Use sample data
        sample_data = report.sample_data
        actual_n = sample_n if sample_n != -1 else sample_data.n_total
        actual_n = min(actual_n, sample_data.n_sample)

        X_arr = sample_data.X_sample[:actual_n]
        y_arr = sample_data.y_sample[:actual_n]
        y_pred = sample_data.y_pred_sample[:actual_n]

        n_x_cols = X_arr.shape[1]
        if sample_data.x_column_names and len(sample_data.x_column_names) == n_x_cols:
            x_col_names = sample_data.x_column_names
        else:
            x_col_names = [f'X{i+1}' for i in range(n_x_cols)]

        data_source = f"sample ({actual_n} of {sample_data.n_total})"
    else:
        return ""

    # Prepare data as JSON for JavaScript
    data_dict = {}
    for i, name in enumerate(x_col_names):
        data_dict[name] = X_arr[:, i].tolist()

    data_dict['Y (Actual)'] = y_arr.tolist()
    data_dict['Y (Predicted)'] = y_pred.tolist()
    data_dict['Residual'] = (y_arr - y_pred).tolist()
    data_dict['Row Index'] = list(range(1, actual_n + 1))

    # All available variables for selection (X columns + Y + derived)
    all_vars = list(x_col_names) + ['Y (Actual)', 'Y (Predicted)', 'Residual', 'Row Index']

    html_parts = []

    html_parts.append("<h2>Interactive Data Explorer</h2>")
    html_parts.append(f"<p>Explore relationships in the data ({actual_n} observations from {data_source}). Select variables for axes, color, and size.</p>")

    # Control panel - all controls in one row
    html_parts.append('<div class="controls">')

    # X variable selector
    html_parts.append('<div class="control-group">')
    html_parts.append('<label for="x-var">X Variable</label>')
    html_parts.append('<select id="x-var" onchange="updateScatterPlot()">')
    for i, var in enumerate(all_vars):
        selected = 'selected' if i == 0 else ''
        html_parts.append(f'<option value="{var}" {selected}>{var}</option>')
    html_parts.append('</select>')
    html_parts.append('</div>')

    # Y variable selector
    html_parts.append('<div class="control-group">')
    html_parts.append('<label for="y-var">Y Variable</label>')
    html_parts.append('<select id="y-var" onchange="updateScatterPlot()">')
    for var in all_vars:
        selected = 'selected' if var == 'Y (Actual)' else ''
        html_parts.append(f'<option value="{var}" {selected}>{var}</option>')
    html_parts.append('</select>')
    html_parts.append('</div>')

    # Color variable selector
    html_parts.append('<div class="control-group">')
    html_parts.append('<label for="color-var">Color By</label>')
    html_parts.append('<select id="color-var" onchange="updateScatterPlot()">')
    html_parts.append('<option value="none" selected>None (single color)</option>')
    for var in all_vars:
        html_parts.append(f'<option value="{var}">{var}</option>')
    html_parts.append('</select>')
    html_parts.append('</div>')

    # Size variable selector
    html_parts.append('<div class="control-group">')
    html_parts.append('<label for="size-var">Size By</label>')
    html_parts.append('<select id="size-var" onchange="updateScatterPlot()">')
    html_parts.append('<option value="none" selected>None (uniform size)</option>')
    for var in all_vars:
        html_parts.append(f'<option value="{var}">{var}</option>')
    html_parts.append('</select>')
    html_parts.append('</div>')

    # Y-axis min control (only control for axis, others removed)
    html_parts.append('<div class="control-group">')
    html_parts.append('<label for="y-min">Y Min</label>')
    html_parts.append('<select id="y-min" onchange="updateScatterPlot()">')
    html_parts.append('<option value="0" selected>0</option>')
    html_parts.append('<option value="auto">Auto</option>')
    html_parts.append('</select>')
    html_parts.append('</div>')

    html_parts.append('</div>')  # end controls

    # Plot container
    html_parts.append('<div id="interactive-scatter" class="plotly-container" style="height: 500px;"></div>')

    # JavaScript for interactive plotting
    html_parts.append(f"""
    <script>
        // Data store
        var plotData = {json.dumps(data_dict)};

        // Normalize array for sizing (scale to 5-30 range)
        function normalizeForSize(arr) {{
            var min = Math.min(...arr);
            var max = Math.max(...arr);
            if (max === min) return arr.map(() => 12);
            return arr.map(v => 5 + 25 * (v - min) / (max - min));
        }}

        function updateScatterPlot() {{
            var xVar = document.getElementById('x-var').value;
            var yVar = document.getElementById('y-var').value;
            var colorVar = document.getElementById('color-var').value;
            var sizeVar = document.getElementById('size-var').value;
            var yMinSel = document.getElementById('y-min').value;

            var x = plotData[xVar];
            var y = plotData[yVar];

            var trace = {{
                type: 'scatter',
                mode: 'markers',
                x: x,
                y: y,
                text: plotData['Row Index'].map((idx, i) => {{
                    var lines = ['Row: ' + idx];
                    lines.push(xVar + ': ' + (typeof x[i] === 'number' ? x[i].toFixed(4) : x[i]));
                    lines.push(yVar + ': ' + (typeof y[i] === 'number' ? y[i].toFixed(4) : y[i]));
                    if (colorVar !== 'none') {{
                        var cv = plotData[colorVar][i];
                        lines.push(colorVar + ': ' + (typeof cv === 'number' ? cv.toFixed(4) : cv));
                    }}
                    if (sizeVar !== 'none') {{
                        var sv = plotData[sizeVar][i];
                        lines.push(sizeVar + ': ' + (typeof sv === 'number' ? sv.toFixed(4) : sv));
                    }}
                    return lines.join('<br>');
                }}),
                hoverinfo: 'text',
                marker: {{
                    size: 10,
                    opacity: 0.7,
                    colorscale: 'Viridis',
                    showscale: false
                }}
            }};

            // Apply color if selected
            if (colorVar !== 'none') {{
                trace.marker.color = plotData[colorVar];
                trace.marker.colorscale = 'Viridis';
                trace.marker.showscale = true;
                trace.marker.colorbar = {{title: colorVar}};
            }} else {{
                trace.marker.color = '#3498db';
            }}

            // Apply size if selected
            if (sizeVar !== 'none') {{
                trace.marker.size = normalizeForSize(plotData[sizeVar]);
                trace.marker.sizemode = 'diameter';
            }}

            // Build axis range settings
            var yAxisConfig = {{title: yVar}};
            var xAxisConfig = {{title: xVar}};

            // Y-axis range: 0 or Auto
            if (yMinSel === '0') {{
                var yMax = Math.max(...y) * 1.05;
                yAxisConfig.range = [0, yMax];
            }}

            var layout = {{
                title: yVar + ' vs ' + xVar,
                xaxis: xAxisConfig,
                yaxis: yAxisConfig,
                hovermode: 'closest',
                showlegend: false
            }};

            Plotly.newPlot('interactive-scatter', [trace], layout, {{responsive: true}});
        }}

        // Initial plot
        updateScatterPlot();
    </script>
    """)

    return "\n".join(html_parts)
